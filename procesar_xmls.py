import sys
import xml.etree.ElementTree as ET
import os
import openpyxl

def extraer_datos_xml(xml_path):
    """
    Extrae el UUID de un archivo XML.
    Retorna el UUID si lo encuentra, de lo contrario None.
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        uuid = None

        # Buscar UUID en cualquier etiqueta
        for elem in root.iter():
            if 'UUID' in elem.attrib:
                uuid = elem.attrib['UUID']
                break  # Solo queremos el primero que encontremos
        
        return uuid

    except ET.ParseError as e:
        print(f"Error al parsear el XML '{os.path.basename(xml_path)}': {e}")
        return None
    except Exception as e:
        print(f"Ocurrió un error inesperado al procesar '{os.path.basename(xml_path)}': {e}")
        return None

def procesar_carpeta_xml(carpeta_xml, archivo_excel_salida):
    """
    Procesa todos los archivos XML en la carpeta especificada,
    extrae sus UUIDs y los guarda en un archivo Excel.
    """
    if not os.path.isdir(carpeta_xml):
        print(f"Error: La carpeta '{carpeta_xml}' no existe.")
        return

    # Crear un nuevo libro de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "UUIDs Extraídos"

    # Escribir los encabezados en la primera fila
    sheet['A1'] = "Nombre del Archivo XML"
    sheet['B1'] = "UUID"

    fila_actual = 2 # Empezamos a escribir datos desde la fila 2

    print(f"Procesando archivos XML en: {carpeta_xml}")

    for filename in os.listdir(carpeta_xml):
        if filename.endswith(".xml"):
            xml_path = os.path.join(carpeta_xml, filename)
            print(f"Extrayendo datos de: {filename}")
            uuid = extraer_datos_xml(xml_path)
            
            # Escribir el nombre del archivo y el UUID en el Excel
            sheet.cell(row=fila_actual, column=1, value=filename)
            sheet.cell(row=fila_actual, column=2, value=uuid if uuid else "No encontrado")
            fila_actual += 1

    try:
        workbook.save(archivo_excel_salida)
        print(f"\nProceso completado. Resultados guardados en: {archivo_excel_salida}")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")

if __name__ == "__main__":
    # Puedes modificar esta ruta a la carpeta donde están tus archivos XML
    # Por ejemplo, si están en la misma carpeta que el script:
    # carpeta_a_procesar = os.path.dirname(os.path.abspath(__file__)) 
    
    # O una ruta específica:
    carpeta_a_procesar = "/Users/danielmendoza/Downloads/" # Modifica esta línea con la ruta de tu carpeta de XMLs
    
    archivo_salida_excel = "UUIDs_extraidos.xlsx" # Nombre del archivo Excel de salida

    # Verificar si se pasó un argumento (aunque ahora no es obligatorio, por si quieres mantener la opción)
    if len(sys.argv) > 1:
        # Si se pasa un argumento, asumimos que es la carpeta a procesar
        carpeta_a_procesar = sys.argv[1]

    procesar_carpeta_xml(carpeta_a_procesar, archivo_salida_excel)
